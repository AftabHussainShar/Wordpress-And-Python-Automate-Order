from woocommerce import API
import openpyxl

# Initialize WooCommerce API
wcdata = API(
    url='https://mikeelford.com/',
    consumer_key='ck_159cf0c676f2012cda4777ed31a58da9b3b54aab',
    consumer_secret='cs_54effa9e46e3c83dbf53bb04637f1f7a130995df',
    version='wc/v3'
)

# Load existing Excel file
wb = openpyxl.load_workbook('Draftpics.xlsx')

# Retrieve or create the specific sheet
if 'From Woocommerce' in wb.sheetnames:
    ws = wb['From Woocommerce']
else:
    ws = wb.create_sheet(title='From Woocommerce')

# Clear existing data from the sheet, if any
ws.delete_rows(2, ws.max_row)

# Retrieve all orders from WooCommerce API
all_orders = []

page = 1
while True:
    try:
        orders_response = wcdata.get('orders', params={'page': page})
        orders = orders_response.json()
        all_orders.extend(orders)
        if 'next' in orders_response.headers.get('Link', ''):
            page += 1
        else:
            break
    except Exception as e:
        print("Error fetching orders:", e)
        break

# Extract relevant fields from WooCommerce API response
for idx, order in enumerate(all_orders, start=2):  # Start from row 2, assuming headers are in row 1
    ws[f'A{idx}'] = order['id']
    ws[f'B{idx}'] = order['date_created']
    if order['status'] == 'custom-status':
        ws[f'C{idx}'] = 'created'
    else:
        ws[f'C{idx}'] = order['status']
    # Add dollar formatting to the "Total" field
    cell = ws[f'E{idx}']
    # cell.number_format = '$#,##0.00'
    # Assign the value to the cell
    cell.value = order['total']
    ws[f'E{idx}'] = next((meta['value'] for meta in order['meta_data'] if meta['key'] == '_billing_select_one'), None)
    ws[f'F{idx}'] = order['billing']['last_name'] + ' ' + order['billing']['first_name']

# Save the modified Excel file
wb.save('Draftpics.xlsx')
print("Data Updtated to Draftpics.xlsx")